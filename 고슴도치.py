import discord
import asyncio
import datetime
import time
import openpyxl
import random



client=discord.Client()


token=("NzgwMDgxMTI5NTU4Mzc2NDcw.X7p5Kg.h6QQhmd6yt3BkulFZVrF2eWnmHE")


@client.event
async def on_ready():
    print(str(client.user.name),"봇이 실행되었습니다!")
    game = discord.Game("가시 손질")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):
    if message.author.bot:
        return None



#명령어


    if message.content.startswith("~명령어"):
        await message.channel.send("~ + ,[명령어, 안녕, 정보, 홀짝, 송금]")



#메이플

    if message.content.startswith("메"):
        file=openpyxl.load_workbook("고슴도치.xlsx")
        sheet=file.active
        for i in range(1,50):
            if sheet["A"+str(i)].value==message.author.id:
                sheet["H"+str(i)].value=sheet["H"+str(i)].value+1
                await message.channel.send(message.author.mention+", 금지 단어가 포함 되어있습니다. 맞스택이 "+str(sheet["H"+str(i)].value)+"적립됩니다.")
                file.save("고슴도치.xlsx")
                break




            if sheet["A"+str(i)].value==None:
                sheet["A"+str(i)].value=message.author.id
                sheet["B"+str(i)].value=message.author.name
                await message.channel.send(message.author.mention+", 등록되었습니다")
                sheet["C"+str(i)].value=10000
                sheet["E"+str(i)].value=0
                sheet["F"+str(i)].value=0
                sheet["H"+str(i)].value=1
                file.save("고슴도치.xlsx")

                break

            else:
                None







 #인사

    if message.content.startswith("~안녕"):
        ment=['"맞?"','"맞고싶은거 같은데"','"정신 나갈거가테~"','"안녕"','"에잇*팔"','"선 넘네?"','"맞고 싶어.?"','"닥*ㅎ"','"나중에 대화하자?"','"점심나가서 먹어?"',"맞고싶음 3스택 적립. "]
        mention=random.choice(ment)
        await message.channel.send(message.author.mention+", "+mention)
        




#정보



    if message.content.startswith("~정보"):
        file=openpyxl.load_workbook("고슴도치.xlsx")
        sheet=file.active
        for i in range(1,50):
            if sheet["A"+str(i)].value==message.author.id:
                embed=discord.Embed(title=message.author.name+"님의 정보",
                                    description=message.author.mention+"\n코인: `"+str(sheet["C"+str(i)].value)+"원`, 전적: `"+str(sheet["E"+str(i)].value)+"승` ,`"+str(sheet["F"+str(i)].value)+"패`",
                                    color=0xDDEED)
                await message.channel.send(embed=embed)
                break




            if sheet["A"+str(i)].value==None:
                sheet["A"+str(i)].value=message.author.id
                sheet["B"+str(i)].value=message.author.name
                await message.channel.send(message.author.mention+", 등록되었습니다")
                sheet["C"+str(i)].value=10000
                sheet["E"+str(i)].value=0
                sheet["F"+str(i)].value=0
                sheet["H"+str(i)].value=0
                file.save("고슴도치.xlsx")

                break

            else:
                None






 #홀짝


    if message.content.startswith("~홀짝"):
        file=openpyxl.load_workbook("고슴도치.xlsx")
        choice=message.content.split(" ")
        sheet=file.active


        luck=["홀","짝"]
        luck_choice=random.choice(luck)
        print(luck_choice)

        if choice[1]==luck_choice:
            money=int(int(choice[2])*1.75)
            a="+"
            b="승"
            c="성공!!"
            

        elif choice[1]!=luck_choice:
            money=-int(choice[2])
            a=""
            b="패"
            c="실패ㅋㅋ"

        
        for i in range(1,50):
            if sheet["A"+str(i)].value==message.author.id:

                if int(choice[2])>sheet["C"+str(i)].value:
                    await message.channel.send(message.author.mention+", 금액이 부족합니다.")
                    break
                
                elif int(choice[2])<=sheet["C"+str(i)].value:
                    sheet["C"+str(i)].value=sheet["C"+str(i)].value+money
                    if b=="승":
                        sheet["E"+str(i)].value=sheet["E"+str(i)].value+1
                    elif b=="패":
                        sheet["F"+str(i)].value=sheet["F"+str(i)].value+1
                    
                    embed=discord.Embed(title=c,
                                        description=message.author.mention+"\n`"+a+str(money)+"`원, 현재 재산: `"+str(sheet["C"+str(i)].value)+"`원",
                                        color=0xDDEED)

                    await message.channel.send(embed=embed)

                    
                    file.save("고슴도치.xlsx")
                    break




        else:
            await message.channel.send(message.author.mention+", 먼저 정보를 등록해주세요.")





 # 송금
 



    if message.content.startswith("~송금"):
        file=openpyxl.load_workbook("고슴도치.xlsx")
        send=message.content.split(" ")
        sheet=file.active
        
        for i in range(1,50):
            if sheet["A"+str(i)].value==message.author.id:

                if int(send[2])>sheet["C"+str(i)].value and int(send[2])>0:
                    await message.channel.send(message.author.mention+", 금액이 부족합니다.")
                    break
                
                elif int(send[2])<=sheet["C"+str(i)].value:
                    sheet["C"+str(i)].value=sheet["C"+str(i)].value-int(send[2])

                    for o in range(1,50):
                        if sheet["B"+str(o)].value==send[1]:
                           sheet["C"+str(o)].value=sheet["C"+str(o)].value+int(send[2])
                           break


                    await message.channel.send(message.author.mention+"님이 "+send[1]+"님에게 `"+send[2]+"`원을 보냈습니다.")
                    
                    embed=discord.Embed(title="송금",
                                        description=message.author.mention+"\n현재 재산: `"+str(sheet["C"+str(i)].value)+"`원",
                                        color=0xDDEED)

                    await message.channel.send(embed=embed)

                    
                    file.save("고슴도치.xlsx")
                    break




        else:
                await message.channel.send(message.author.mention+", 먼저 정보를 등록해주세요.")





#강화












client.run(token)
